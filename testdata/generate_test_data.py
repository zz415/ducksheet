"""
DuckSheet test data generator.
Produces testdata.xlsx with three sheets:
  - sales      (~500 rows)  mixed types: date, string, int, float, bool, some nulls
  - products   (~50 rows)   product catalog
  - large      (~5000 rows) scale test
"""

import random
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(42)

# ── helpers ──────────────────────────────────────────────────────────────────

def rand_date(start_year=2022, end_year=2026):
    start = datetime.date(start_year, 1, 1)
    end   = datetime.date(end_year, 3, 1)
    return start + datetime.timedelta(days=random.randint(0, (end - start).days))

def maybe_null(value, pct=0.05):
    return None if random.random() < pct else value

def style_header_row(ws, n_cols):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

# ── sheet: sales ─────────────────────────────────────────────────────────────

REGIONS     = ["North", "South", "East", "West", "Central"]
CATEGORIES  = ["Electronics", "Furniture", "Office Supplies", "Software", "Services"]
REPS        = ["Alice Johnson", "Bob Martinez", "Carol Lee", "David Kim",
               "Eve Torres", "Frank Nguyen", "Grace Patel", "Henry Wu"]
STATUSES    = ["Closed Won", "Closed Lost", "Pending", "In Review"]

def build_sales(ws):
    headers = [
        "order_id", "order_date", "close_date", "rep_name", "region",
        "category", "product_name", "quantity", "unit_price",
        "discount_pct", "revenue", "is_renewal", "status", "notes"
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))

    products = {
        "Electronics":    ["Laptop Pro 15", "Wireless Headset", "USB-C Hub", "Monitor 27\""],
        "Furniture":      ["Standing Desk", "Ergonomic Chair", "Bookshelf Unit", "Filing Cabinet"],
        "Office Supplies":["Printer Paper (case)", "Stapler Set", "Whiteboard Kit", "Label Maker"],
        "Software":       ["Analytics Suite", "Security Bundle", "CRM License", "ERP Module"],
        "Services":       ["Onboarding Package", "Annual Support", "Training Day", "Consulting Block"],
    }

    for i in range(1, 501):
        order_date   = rand_date()
        close_date   = order_date + datetime.timedelta(days=random.randint(1, 90))
        category     = random.choice(CATEGORIES)
        product      = random.choice(products[category])
        qty          = random.randint(1, 50)
        unit_price   = round(random.uniform(19.99, 4999.99), 2)
        discount     = maybe_null(round(random.uniform(0, 0.30), 2), pct=0.15)
        disc_val     = discount if discount is not None else 0
        revenue      = round(qty * unit_price * (1 - disc_val), 2)
        is_renewal   = random.choice([True, False])
        status       = random.choice(STATUSES)
        notes        = maybe_null(random.choice([
            "Expedited shipping requested", "Multi-year contract",
            "Pending legal review", "VIP account", "Referral deal",
            "Bundled with support", None, None, None   # weighted toward None
        ]), pct=0.0)  # nulls already baked in via the list

        ws.append([
            f"ORD-{i:05d}",
            order_date,
            maybe_null(close_date, pct=0.08),
            random.choice(REPS),
            random.choice(REGIONS),
            category,
            product,
            qty,
            unit_price,
            discount,
            revenue,
            is_renewal,
            status,
            notes,
        ])

    # format date columns
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "MM/DD/YYYY"

    auto_width(ws)

# ── sheet: products ───────────────────────────────────────────────────────────

SUPPLIERS = ["Apex Supply Co.", "Delta Wholesale", "Prime Distributors",
             "Summit Goods", "Horizon Logistics"]

def build_products(ws):
    headers = [
        "product_id", "sku", "product_name", "category", "supplier",
        "cost_price", "list_price", "stock_qty", "reorder_level",
        "is_active", "last_restocked", "weight_kg", "notes"
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))

    names = [
        ("Electronics",    "Laptop Pro 15"),   ("Electronics",    "Wireless Headset"),
        ("Electronics",    "USB-C Hub"),        ("Electronics",    'Monitor 27"'),
        ("Electronics",    "Mechanical Keyboard"), ("Electronics", "Webcam 4K"),
        ("Furniture",      "Standing Desk"),    ("Furniture",      "Ergonomic Chair"),
        ("Furniture",      "Bookshelf Unit"),   ("Furniture",      "Filing Cabinet"),
        ("Furniture",      "Visitor Chair"),    ("Furniture",      "Meeting Table"),
        ("Office Supplies","Printer Paper (case)"), ("Office Supplies","Stapler Set"),
        ("Office Supplies","Whiteboard Kit"),   ("Office Supplies","Label Maker"),
        ("Office Supplies","Ballpoint Pens (box)"), ("Office Supplies","Folder Pack 50"),
        ("Software",       "Analytics Suite"),  ("Software",       "Security Bundle"),
        ("Software",       "CRM License"),      ("Software",       "ERP Module"),
        ("Software",       "Backup Service"),   ("Software",       "SSO Add-on"),
        ("Services",       "Onboarding Package"), ("Services",     "Annual Support"),
        ("Services",       "Training Day"),     ("Services",       "Consulting Block"),
        ("Services",       "Data Migration"),   ("Services",       "Health Check"),
    ]

    for i, (cat, name) in enumerate(names, start=1):
        cost       = round(random.uniform(5, 1500), 2)
        list_price = round(cost * random.uniform(1.3, 2.5), 2)
        stock      = maybe_null(random.randint(0, 500), pct=0.06)
        reorder    = random.randint(5, 50)
        is_active  = random.choice([True, True, True, False])
        restocked  = maybe_null(rand_date(2024, 2026), pct=0.10)
        weight     = maybe_null(round(random.uniform(0.1, 25.0), 2), pct=0.20)

        ws.append([
            i,
            f"SKU-{cat[:3].upper()}-{i:03d}",
            name,
            cat,
            random.choice(SUPPLIERS),
            cost,
            list_price,
            stock,
            reorder,
            is_active,
            restocked,
            weight,
            maybe_null("Discontinued soon", pct=0.85),
        ])

    for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "MM/DD/YYYY"

    auto_width(ws)

# ── sheet: large ──────────────────────────────────────────────────────────────

DEPARTMENTS  = ["Engineering", "Marketing", "Sales", "Finance", "HR",
                "Operations", "Legal", "Product", "Design", "Support"]
JOB_TITLES   = ["Analyst", "Senior Analyst", "Manager", "Director",
                "Coordinator", "Specialist", "Associate", "Lead"]
FIRST_NAMES  = ["James","Mary","John","Patricia","Robert","Jennifer","Michael",
                "Linda","William","Barbara","David","Susan","Richard","Jessica",
                "Joseph","Sarah","Thomas","Karen","Charles","Lisa"]
LAST_NAMES   = ["Smith","Johnson","Williams","Brown","Jones","Garcia","Miller",
                "Davis","Rodriguez","Martinez","Hernandez","Lopez","Gonzalez",
                "Wilson","Anderson","Thomas","Taylor","Moore","Jackson","Martin"]
CITIES       = ["New York","Los Angeles","Chicago","Houston","Phoenix",
                "Philadelphia","San Antonio","San Diego","Dallas","San Jose",
                "Austin","Jacksonville","Fort Worth","Columbus","Charlotte"]

def build_large(ws):
    headers = [
        "employee_id", "first_name", "last_name", "email", "department",
        "job_title", "hire_date", "salary", "bonus_pct", "is_fulltime",
        "city", "years_exp", "performance_score", "last_review_date", "manager_id"
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))

    for i in range(1, 5001):
        first    = random.choice(FIRST_NAMES)
        last     = random.choice(LAST_NAMES)
        hire     = rand_date(2010, 2026)
        salary   = round(random.uniform(38000, 210000), 2)
        bonus    = maybe_null(round(random.uniform(0.02, 0.20), 2), pct=0.20)
        score    = maybe_null(round(random.uniform(1.0, 5.0), 1), pct=0.05)
        review   = maybe_null(rand_date(2023, 2026), pct=0.10)
        mgr_id   = maybe_null(random.randint(1, 500), pct=0.05)

        ws.append([
            i,
            first,
            last,
            f"{first.lower()}.{last.lower()}{i}@company.com",
            random.choice(DEPARTMENTS),
            random.choice(JOB_TITLES),
            hire,
            salary,
            bonus,
            random.choice([True, True, True, False]),
            random.choice(CITIES),
            random.randint(0, 25),
            score,
            review,
            mgr_id,
        ])

    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "MM/DD/YYYY"
    for row in ws.iter_rows(min_row=2, min_col=14, max_col=14):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "MM/DD/YYYY"

    auto_width(ws)

# ── main ──────────────────────────────────────────────────────────────────────

wb = Workbook()

ws_sales = wb.active
ws_sales.title = "sales"
build_sales(ws_sales)

ws_products = wb.create_sheet("products")
build_products(ws_products)

ws_large = wb.create_sheet("large")
build_large(ws_large)

out_path = r"c:\claude_projects\ducksheet\testdata\testdata.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"  sales:    500 rows, 14 cols")
print(f"  products:  30 rows, 13 cols")
print(f"  large:   5000 rows, 15 cols")
